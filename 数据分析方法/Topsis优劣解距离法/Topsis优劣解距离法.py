import numpy as np
import pandas as pd
import openpyxl


# 从excel文件中读取数据
def read(file):
    # 打开文件
    wb = openpyxl.load_workbook(filename=file)
    # 通过索引获取表格
    sheet = wb["Sheet1"]
    # 获取行数
    rows = sheet.max_row
    # 存放读取的数据
    all_content = []
    # 取第2~第5列的数据
    for j in range(2, 6):
        temp = []
        for i in range(2, rows + 1):
            # 获取数据
            cell = int(sheet.cell(row=i, column=j).value)
            temp.append(cell)
        # 按列添加到结果集中
        all_content.append(temp)
        temp = []
    return np.array(all_content)


# 极小型指标 -> 极大型指标
def convert_1(datas):
    return np.max(datas) - datas


# 中间型指标 -> 极大型指标
def convert_2(datas, x_best):
    temp_datas = datas - x_best
    M = np.max(abs(temp_datas))
    answer_datas = 1 - abs(datas - x_best) / M
    return answer_datas


# 区间型指标 -> 极大型指标
def convert_3(datas, x_min, x_max):
    M = max(x_min - np.min(datas), np.max(datas) - x_max)
    answer_list = []
    for i in datas:
        if i < x_min:
            answer_list.append(1 - (x_min - i) / M)
        elif x_min <= i <= x_max:
            answer_list.append(1)
        else:
            answer_list.append(1 - (i - x_max) / M)
    return np.array(answer_list)


# 正向化矩阵标准化
def matrixTrans_1(datas):
    K = np.power(np.sum(pow(datas, 2), axis=1), 0.5)
    for i in range(0, K.size):
        for j in range(0, datas[i].size):
            # 套用矩阵标准化的公式
            datas[i, j] = datas[i, j] / K[i]
    return datas


# 计算得分并归一化
def matrixTrans_2(answer2):
    list_max = np.array(
        [np.max(answer2[0, :]), np.max(answer2[1, :]), np.max(answer2[2, :]), np.max(answer2[3, :])])  # 获取每一列的最大值
    list_min = np.array(
        [np.min(answer2[0, :]), np.min(answer2[1, :]), np.min(answer2[2, :]), np.min(answer2[3, :])])  # 获取每一列的最小值
    max_list = []  # 存放第i个评价对象与最大值的距离
    min_list = []  # 存放第i个评价对象与最小值的距离
    answer_list = []  # 存放评价对象的未归一化得分
    for k in range(0, np.size(answer2, axis=1)):  # 遍历每一列数据
        max_sum = 0
        min_sum = 0
        for q in range(0, 4):  # 有四个指标
            max_sum += np.power(answer2[q, k] - list_max[q], 2)  # 按每一列计算Di+
            min_sum += np.power(answer2[q, k] - list_min[q], 2)  # 按每一列计算Di-
        max_list.append(pow(max_sum, 0.5))
        min_list.append(pow(min_sum, 0.5))
        answer_list.append(min_list[k] / (min_list[k] + max_list[k]))  # 套用计算得分的公式 Si = (Di-) / ((Di+) +(Di-))
        max_sum = 0
        min_sum = 0
    answer = np.array(answer_list)  # 得分归一化
    return answer / np.sum(answer)


def main():
    file = "/Users/mark/建模课程/数据分析方法/Topsis优劣解距离法/river.xlsx"
    # 读取文件
    answer1 = read(file)
    answer2 = []
    # 按照不同的列，根据不同的指标转换为极大型指标，因为只有四列
    for i in range(0, 4):
        answer = None
        # 本来就是极大型指标，不用转换
        if i == 0:
            answer = answer1[0]
        # 中间型指标
        elif i == 1:
            answer = convert_2(answer1[1], 7)
        # 极小型指标
        elif i == 2:
            answer = convert_1(answer1[2])
        # 范围型指标
        else:
            answer = convert_3(answer1[3], 10, 20)
        answer2.append(answer)
    # 将list转换为numpy数组
    answer2 = np.array(answer2)
    # 数组正向化
    answer3 = matrixTrans_1(answer2)
    # 标准化处理去量纲
    answer4 = matrixTrans_2(answer3)
    # 计算得分
    data = pd.DataFrame(answer4)

    # 将得分输出到excel表格中
    writer = pd.ExcelWriter("/Users/mark/建模课程/数据分析方法/Topsis优劣解距离法/result.xlsx")  # 写入Excel文件
    data.to_excel(writer, 'page_1', float_format='%.5f')  # ‘page_1’是写入excel的sheet名
    writer.save()
    writer.close()


main()
